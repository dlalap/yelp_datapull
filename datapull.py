# Yelpscrape
# by Dean Lalap
# (c) 2017

import urllib
import re
import openpyxl
import os
import sentiment

from bs4 import BeautifulSoup
from threading import Thread
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sentiment import analyze, print_result
from textClassify import classify, index

#########################################
############## USER INPUT ###############
#########################################

mainUrl = 'HTTP://YELP.COM/CLIENT_URL_HERE'
pagesToScan = 1

#########################################
############ END USER INPUT #############
#########################################

wb = Workbook()
resultsWorkbook = load_workbook('yelpDataPullTest.xlsx')
resultsWorksheet = resultsWorkbook.get_sheet_names()[0]
resultsList = resultsWorkbook[resultsWorksheet]

# Grab the active worksheet
ws = wb.active

reviewTextCounter = 1
ratingValueCounter = 1
dateCounter = 1

reviewResults = [u'Review Text']
ratingResults = [u'Rating']
dateResults = [u'Date']

sentimentResults = {}
classificationResults = {}

url = []
for pages in range(pagesToScan):
	startVal = 20*pages
	url.append(mainUrl + '?start=' + str(startVal) + '&sort_by=date_desc')

def loadPage(ur):
	html = urllib.urlopen(eachLink).read()
	content = BeautifulSoup(html, 'html.parser')
	return content

def scrapeAll(content):
	# content = loadPage(ur)
	# i = 0
	# for review in content.select('p'):
	print(content.prettify())


def pullReviewText(content):
	# content = loadPage(ur)
	resultsToAppend = []
	for review in content.select('p[lang]'):

		print("\n\n")
		global reviewTextCounter
		print(reviewTextCounter)
		reviewTextCounter += 1
		print(''.join(review(text=True)))
		reviewText = ''.join(review(text=True))
		print("\n\n")
		resultsToAppend.append(reviewText)
		print resultsToAppend
	return resultsToAppend

def pullRating(content):
	# content = loadPage(ur)
	resultsToAppend = []
	for rating in content.findAll("meta", {"itemprop":"ratingValue"})[1:]:
		print("\n\n")
		global ratingValueCounter
		print(ratingValueCounter)
		print(rating.get('content','')) 
		ratingValue = rating.get('content', '')
		ratingValueCounter += 1
		resultsToAppend.append(ratingValue)
	return resultsToAppend

def pullDate(content):
	resultsToAppend = []
	# content = loadPage(ur)
	for date in content.findAll("meta", {"itemprop":"datePublished"}):
		print("\n\n")
		global dateCounter
		print(dateCounter)
		print(date.get('content',''))
		dateValue = date.get('content','')
		dateCounter += 1
		resultsToAppend.append(dateValue)
	return resultsToAppend

for eachLink in url:
	site = loadPage(url)

	reviewResults.extend(pullReviewText(site))
	ratingResults.extend(pullRating(site))
	dateResults.extend(pullDate(site))

print(reviewResults)
print(ratingResults)
print(dateResults)

for item in range(len(dateResults)):
	dateIndex = 'A' + str(item+1)
	reviewIndex = 'B' + str(item+1)
	ratingIndex = 'C' + str(item+1)
	resultsList[dateIndex] = dateResults[item]
	resultsList[reviewIndex] = reviewResults[item]
	resultsList[ratingIndex] = ratingResults[item]

resultsWorkbook.save('yelpDataPullTest_output.xlsx')