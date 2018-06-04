# Yelpscrape
# by Dean Lalap
# (c) 2017

import urllib
import re
import openpyxl
import os
import sentiment
import google

from bs4 import BeautifulSoup
from threading import Thread
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, colors, Color
from sentiment import *
from textClassify import *

#########################################
############## USER INPUT ###############
#########################################

mainUrl = 'https://www.yelp.com/biz/mr-chow-new-york'
pagesToScan = 14

#########################################
############ END USER INPUT #############
#########################################

wb = Workbook()
resultsWorkbook = load_workbook('yelpDataPullTest.xlsx')
resultsWorksheet = resultsWorkbook.get_sheet_names()[0]
resultsList = resultsWorkbook[resultsWorksheet]

# Grab the active worksheet
ws = wb.active

reviewTextCounter = 1
ratingValueCounter = 1
dateCounter = 1

reviewResults = [u'Review Text']
ratingResults = [u'Rating']
dateResults = [u'Date']

sentimentResults = {1:'Sentiment'}
classificationResults = {1: 'Classify'}

url = []
for pages in range(pagesToScan):
	startVal = 20*pages
	url.append(mainUrl + '?start=' + str(startVal) + '&sort_by=date_desc')

def loadPage(ur):
	html = urllib.urlopen(eachLink).read()
	content = BeautifulSoup(html, 'html.parser')
	return content

def scrapeAll(content):
	# content = loadPage(ur)
	# i = 0
	# for review in content.select('p'):
	print(content.prettify())


def pullReviewText(content):
	# content = loadPage(ur)
	resultsToAppend = []
	reviewSentiments = {}
	reviewClassify = {}

	for review in content.select('p[lang]'):

		# print("\n\n")
		global reviewTextCounter
		# print(reviewTextCounter)
		reviewTextCounter += 1
		# print(''.join(review(text=True)))
		reviewText = ''.join(review(text=True))
		print(reviewText)
		reviewSentiments[reviewTextCounter] = analyze(reviewText)
		try:
			reviewClassify[reviewTextCounter] = classify(reviewText)
		except google.api_core.exceptions.ClientError:
			reviewClassify[reviewTextCounter] = 'Insufficient amount to classify'
		# print("\n\n")
		resultsToAppend.append(reviewText)
		#print(resultsToAppend)

	return resultsToAppend, reviewSentiments, reviewClassify

def pullRating(content):
	# content = loadPage(ur)
	resultsToAppend = []
	for rating in content.findAll("meta", {"itemprop":"ratingValue"})[1:]:
		global ratingValueCounter
		# print("\n\n")
		# print(ratingValueCounter)
		# print(rating.get('content','')) 
		ratingValue = rating.get('content', '')
		ratingValueCounter += 1
		resultsToAppend.append(ratingValue)
	return resultsToAppend

def pullDate(content):
	resultsToAppend = []
	# content = loadPage(ur)
	for date in content.findAll("meta", {"itemprop":"datePublished"}):
		global dateCounter
		# print("\n\n")
		# print(dateCounter)
		# print(date.get('content',''))
		dateValue = date.get('content','')
		dateCounter += 1
		resultsToAppend.append(dateValue)
	return resultsToAppend

for eachLink in url:
	site = loadPage(url)

	reviewClassifySentiments = pullReviewText(site)

	reviewResults.extend(reviewClassifySentiments[0])
	ratingResults.extend(pullRating(site))
	dateResults.extend(pullDate(site))

	sentimentResults.update(reviewClassifySentiments[1])
	classificationResults.update(reviewClassifySentiments[2])
print('\n\n')
print('sentimentResults = ' + str(sentimentResults))
print('classificationResults = ' + str(classificationResults))
print('\n\n')
# def sentimentCellColor(value):
# 	cellStyle = 0
# 	try:
# 		if value < -0.3:
# 			fontStyle = Font(color=colors.WHITE)
# 			cellStyle = PatternFill(color=colors.RED)
# 		elif value > -0.3 & value < 0:
# 			fontStyle = Font(color=colors.BLACK)
# 			cellStyle = PatternFill(color=colors.ORANGE)
# 		elif value > 0 & value < 0.3:
# 			fontStyle = Font(color=colors.BLACK)
# 			cellStyle = PatternFill(color=colors.YELLOW)
# 		else:
# 			fontStyle = Font(color=colors.BLACK)
# 			cellStyle = PatternFill(color=colors.GREEN)
# 	except TypeError:
# 	return cellStyle

# print(reviewResults)
# print(ratingResults)
# print(dateResults)

for item in range(len(dateResults)):
	dateIndex = 'A' + str(item+1)
	reviewIndex = 'B' + str(item+1)
	ratingIndex = 'C' + str(item+1)
	classifyFirstIndex = 'D' + str(item+1)
	classifySecondIndex = 'E' + str(item+1)
	classifyThirdIndex = 'F' + str(item+1)
	sentimentScoreIndex = 'G' + str(item+1)
	sentimentMagnitudeIndex = 'H' + str(item+1)
	resultsList[dateIndex] = dateResults[item]
	resultsList[reviewIndex] = reviewResults[item]
	resultsList[ratingIndex] = ratingResults[item]
	try:
		resultsList[classifyFirstIndex] = classificationResults[item+1].keys()[0]
	except (AttributeError, IndexError) as e:
		pass
	try:
		resultsList[classifySecondIndex] = classificationResults[item+1].keys()[1]
	except (AttributeError, IndexError) as e:
		pass
	try:
		resultsList[classifyThirdIndex] = classificationResults[item+1].keys()[2]
	except (AttributeError, IndexError) as e:
		pass
	resultsList[sentimentScoreIndex] = sentimentResults[item+1][0]
	# resultsList[sentimentScoreIndex].fill = sentimentCellColor(sentimentResults[item+1][0])
	resultsList[sentimentMagnitudeIndex] = sentimentResults[item+1][1]

	for column in range(len(sentimentResults[item+1][2])):
		resultsList[get_column_letter(column+1+9) + str(item+1)] = sentimentResults[item+1][2][column]
		# resultsList[get_column_letter(column+1+6) + str(item+1)].fill = sentimentCellColor(sentimentResults[item+1][2][column])


resultsWorkbook.save('yelpSentimentTest_output.xlsx')

print('\n\nEnd of file\n\n')